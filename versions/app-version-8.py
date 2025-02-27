import dash
from dash import Dash, dcc, html, Input, Output, State, callback
import os
import pandas as pd
import base64
import io
import openpyxl
import dash_daq as daq
import dash_bootstrap_components as dbc
import plotly.graph_objects as go
import glob
from snirf.create_snirf import create_snirf
import numpy as np
import plotly.graph_objects as go
from concentrations_ucln_srs.ucln_srs import PhantomDataProcessor
from concentrations_ucln_srs.ucln_srs import SRS
from concentrations_ucln_srs.ucln_srs import dual_slope_wavelength
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows  # Import this function
import plotly.express as px
import flask

# Create the Dash app
app = dash.Dash(__name__, suppress_callback_exceptions=True)
server = app.server

app.layout = html.Div([
    # Header Section
    html.Div(
        style={
            'height': '40px',
            'padding': '15px',
            'background': '#ed6a28',
            'color': 'white',
            'display': 'flex',
            'alignItems': 'center',
            'justifyContent': 'center',
            'fontSize': '30px',
            'fontWeight': '100',
            'boxShadow': '0 4px 8px rgba(0, 0, 0, 0.2)',
            'borderBottom': '5px solid #ed6a28'
        },
        children=[
            html.H1(
                'FetalsenseM V2 Dashboard',
                style={'margin': '0', 'margin-left': '10px'}
            )
        ]
    ),
    html.Br(),
    # Flex container for the left and right sections
    html.Div([
        # Left side (1/4th width)
        html.Div([
            # Tabs for General, Data Clean, Data Analysis, and Concentrations
            dcc.Tabs(id='left-tabs', children=[
                dcc.Tab(label='General', children=[
                    # File Upload Section inside the General tab
                    html.Div([
                        html.Div([
                            html.H3('File Upload', style={
                                'background': '#ed6a28',
                                'padding': '15px',
                                'textAlign': 'center',
                                'color': 'white',
                                'fontWeight': 'bold',
                                'fontSize': '48px'
                            }),
                            dcc.Upload(
                                id='upload-data',
                                children=html.Div(['Drag and Drop or ', html.A('Select Files')]),
                                style={'width': '95%', 'height': '70px', 'lineHeight': '70px',
                                       'borderWidth': '1px', 'borderStyle': 'dashed', 'borderRadius': '5px',
                                       'textAlign': 'center', 'margin': '15px', 'fontSize': '24px'},
                                multiple=False,
                            ),
                        ]),
                        html.Div(id='file-names'),
                        dcc.Store(id='store-file-path'),
                        html.Br(),
                        html.Div([
                            html.H3('Download SNIRF', style={
                                'background': '#ed6a28',
                                'padding': '15px',
                                'textAlign': 'center',
                                'color': 'white',
                                'fontWeight': 'bold',
                                'fontSize': '48px'
                            }),
                            html.Button("Download Raw Data SNIRF", id="btn_rawdata_snirf", style={
                                'width': '95%', 'height': '70px', 'lineHeight': '70px',
                                'borderWidth': '1px', 'borderStyle': 'dashed', 'borderRadius': '5px',
                                'textAlign': 'center', 'margin': '15px', 'fontSize': '24px'}),
                            dcc.Download(id="download-file-snirf"),
                        ]),
                        html.Br(),
                        html.Div([
                            html.H3('Resampling Options', style={
                                'textAlign': 'center', 'fontSize': '40px', 'color': '#003f5c'}),
                            dcc.RadioItems(
                                id='resample-option',
                                options=[
                                    {'label': '1Hz Averaging', 'value': 'average'},
                                    {'label': '1Hz Accumulation', 'value': 'accumulation'}
                                ],
                                value=None,  # Default value
                                labelStyle={'display': 'block', 'fontSize': '40px', 'color': '#003f5c'}
                            ),
                        ], style={'textAlign': 'center', 'marginBottom': '20px'}),
                        html.Br(),
                        html.Div([
                            html.H3('View Intensities', style={
                                'background': '#ed6a28',
                                'padding': '15px',
                                'textAlign': 'center',
                                'color': 'white',
                                'fontWeight': 'bold',
                                'fontSize': '48px'
                            }),
                            html.Div([
                                html.H4('Select one or more:', style={
                                    'textAlign': 'left',
                                    'fontSize': '30px',
                                    'marginBottom': '15px',
                                    'color': '#ed6a28',
                                    "font-weight": "100"
                                }),
                                dcc.Dropdown(
                                    id='intensities-options-dropdown',
                                    options=[{'label': option, 'value': option} for option in [
                                        'LED_A_784_DET1', 'LED_A_784_DET2', 'LED_A_784_DET3',
                                        'LED_A_800_DET1', 'LED_A_800_DET2', 'LED_A_800_DET3',
                                        'LED_A_818_DET1', 'LED_A_818_DET2', 'LED_A_818_DET3',
                                        'LED_A_835_DET1', 'LED_A_835_DET2', 'LED_A_835_DET3',
                                        'LED_A_851_DET1', 'LED_A_851_DET2', 'LED_A_851_DET3',
                                        'LED_A_881_DET1', 'LED_A_881_DET2', 'LED_A_881_DET3',
                                        'LED_A_DARK_DET1', 'LED_A_DARK_DET2', 'LED_A_DARK_DET3',
                                        'LED_B_784_DET1', 'LED_B_784_DET2', 'LED_B_784_DET3',
                                        'LED_B_800_DET1', 'LED_B_800_DET2', 'LED_B_800_DET3',
                                        'LED_B_818_DET1', 'LED_B_818_DET2', 'LED_B_818_DET3',
                                        'LED_B_835_DET1', 'LED_B_835_DET2', 'LED_B_835_DET3',
                                        'LED_B_851_DET1', 'LED_B_851_DET2', 'LED_B_851_DET3',
                                        'LED_B_881_DET1', 'LED_B_881_DET2', 'LED_B_881_DET3',
                                        'LED_B_DARK_DET1', 'LED_B_DARK_DET2', 'LED_B_DARK_DET3'
                                    ]],
                                    multi=True,
                                    value=[],
                                    style={'borderColor': '#ed6a28', 'fontSize': '24px'}
                                )
                            ]),
                            html.Div(id='intensity-selection-status', children='Select to view', style={
                                'fontFamily': 'Courier New',
                                'fontSize': '20px',
                                'textAlign': 'center',
                                'marginTop': '15px',
                                'color': '#2B2D42'
                            }),
                            html.Br(),
                            html.Div([
    html.H4('Select Groups', style={
        'textAlign': 'left',
        'fontSize': '30px',
        'marginBottom': '15px',
        'color': '#ed6a28',
        "font-weight": "100"
    }),
    # Group A and Group B selectors placed side by side using flexbox
    html.Div([
        html.Div([
            html.H4('GroupA_Detector1', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
            daq.BooleanSwitch(
                id='groupA_dect1_spectras',
                on=False,
                style={'transform': 'scale(1.1)'}
            )
        ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginRight': '10px'}),

        html.Div([
            html.H4('GroupB_Detector1', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
            daq.BooleanSwitch(
                id='groupB_dect1_spectras',
                on=False,
                style={'transform': 'scale(1.1)'}
            )
        ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginLeft': '10px'})
    ], style={'display': 'flex', 'marginBottom': '10px'}),  # Flexbox for side-by-side
    
    html.Div([
        html.Div([
            html.H4('GroupA_Detector2', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
            daq.BooleanSwitch(
                id='groupA_dect2_spectras',
                on=False,
                style={'transform': 'scale(1.1)'}
            )
        ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginRight': '10px'}),

        html.Div([
            html.H4('GroupB_Detector2', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
            daq.BooleanSwitch(
                id='groupB_dect2_spectras',
                on=False,
                style={'transform': 'scale(1.1)'}
            )
        ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginLeft': '10px'})
    ], style={'display': 'flex', 'marginBottom': '10px'}),

    # Adding GroupB_Detector3 under GroupB_Detector2
    html.Div([
        html.Div([
            html.H4('GroupA_Detector3', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
            daq.BooleanSwitch(
                id='groupA_dect3_spectras',
                on=False,
                style={'transform': 'scale(1.1)'}
            )
        ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginRight': '10px'}),

        html.Div([
            html.H4('GroupB_Detector3', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
            daq.BooleanSwitch(
                id='groupB_dect3_spectras',
                on=False,
                style={'transform': 'scale(1.1)'}
            )
        ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginLeft': '10px'})
    ], style={'display': 'flex', 'marginBottom': '10px'}),

    html.Div([
        html.H4('Select All', style={'fontSize': '20px', 'color': '#cc4c0c', "font-weight": "100"}),
        daq.BooleanSwitch(
            id='select_all_switch',
            on=False,
            style={'transform': 'scale(1.1)', 'padding': '10px'}
        )
    ], style={'display': 'flex', 'alignItems': 'center', 'marginBottom': '10px'}),
                            ]),
                            dbc.Button('View Intensity Over Time ', id='view-graph-btn', color='primary', style={
                                'padding': '15px', 'width': '100%', 'margin': '15px 0', 'fontSize': '22px'}),
                            html.Div(id='select-intensities', children='Select one or multiple groups', style={
                                'fontFamily': 'Courier New',
                                'fontSize': '20px',
                                'textAlign': 'center',
                                'marginTop': '15px',
                                'color': '#2B2D42'
                            }),
                            html.Br(),
html.Div(children=[
    html.H3('Raw Data Quality Check', style={
        'background': '#ed6a28',
        'padding': '15px',
        'textAlign': 'center',
        'color': 'white',
        'fontWeight': 'bold',
        'fontSize': '48px',
    }),
    # Alert and button styles updated for clearer and larger font
    dbc.Alert(
        html.Div([
            html.H4('Signal Noise Ratio', style={'color': '#ed6a28', 'marginLeft': '40px', 'textAlign': 'left', 'fontSize': '28px', 'fontWeight': 'lighter'}),
            html.Button('×', id='snr-close-btn', n_clicks=0, style={'background': 'none', 'border': 'none', 'color': 'black', 'fontSize': '28px', 'cursor': 'pointer', 'float': 'right', 'color': '#ed6a28'})
        ]),
        id='snr-alert',  # Unique ID
        is_open=True,
        dismissable=True,
        style={'marginTop': '15px', 'boxShadow': '0 4px 6px rgba(0, 0, 0, 0.1)'}
    ),
    dbc.Alert(
        html.Div([
            html.H4('Average Signal Noise Ratio', style={'color': '#ed6a28', 'marginLeft': '40px', 'textAlign': 'left', 'fontSize': '28px', 'fontWeight': 'lighter'}),
            html.Button('×', id='avg-snr-close-btn', n_clicks=0, style={'background': 'none', 'border': 'none', 'color': 'black', 'fontSize': '28px', 'cursor': 'pointer', 'float': 'right', 'color': '#ed6a28'})
        ]),
        id='avg-snr-alert',  # Unique ID
        is_open=True,
        dismissable=True,
        style={'marginTop': '15px', 'boxShadow': '0 4px 6px rgba(0, 0, 0, 0.1)'}
    ),
    dbc.Alert(
        html.Div([
            html.H4('Noise Equivalent Power', style={'color': '#ed6a28', 'marginLeft': '40px', 'textAlign': 'left', 'fontSize': '28px', 'fontWeight': 'lighter'}),
            html.Button('×', id='nep-close-btn', n_clicks=0, style={'background': 'none', 'border': 'none', 'color': 'black', 'fontSize': '28px', 'cursor': 'pointer', 'float': 'right', 'color': '#ed6a28'})
        ]),
        id='nep-alert',  # Unique ID
        is_open=True,
        dismissable=True,
        style={'marginTop': '15px', 'boxShadow': '0 4px 6px rgba(0, 0, 0, 0.1)'}
    ),
    dbc.Alert(
        html.Div([
            html.H4('Scatter Plot', style={'color': '#ed6a28', 'marginLeft': '40px', 'textAlign': 'left', 'fontSize': '28px', 'fontWeight': 'lighter'}),
            html.Button('×', id='scatter-plot-btn', n_clicks=0, style={'background': 'none', 'border': 'none', 'color': 'black', 'fontSize': '28px', 'cursor': 'pointer', 'float': 'right', 'color': '#ed6a28'})
        ]),
        id='scatter-plot-alert',  # Unique ID
        is_open=True,
        dismissable=True,
        style={'marginTop': '15px', 'boxShadow': '0 4px 6px rgba(0, 0, 0, 0.1)'}
    ),
    dbc.Alert(
        html.Div([
            html.H4('Distance from Dark', style={'color': '#ed6a28', 'marginLeft': '40px', 'textAlign': 'left', 'fontSize': '28px', 'fontWeight': 'lighter'}),
            html.Button('×', id='distance-from-dark-btn', n_clicks=0, style={'background': 'none', 'border': 'none', 'color': 'black', 'fontSize': '28px', 'cursor': 'pointer', 'float': 'right', 'color': '#ed6a28'})
        ]),
        id='distance-from-dark-alert',  # Unique ID
        is_open=True,
        dismissable=True,
        style={'marginTop': '15px', 'boxShadow': '0 4px 6px rgba(0, 0, 0, 0.1)'}
    ),
    html.Br(),
    html.Div([
        html.H4('Select Groups', style={
            'textAlign': 'left',
            'fontSize': '30px',
            'marginBottom': '15px',
            'color': '#ed6a28',
            "font-weight": "100"
        }),

        # Group A and Group B selectors placed side by side using flexbox
        html.Div([
            html.Div([
                html.H4('GroupA_Detector1', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
                daq.BooleanSwitch(
                    id='groupA_dect1_switch',
                    on=False,
                    style={'transform': 'scale(1.1)'}
                )
            ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginRight': '10px'}),

            html.Div([
                html.H4('GroupB_Detector1', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
                daq.BooleanSwitch(
                    id='groupB_dect1_switch',
                    on=False,
                    style={'transform': 'scale(1.1)'}
                )
            ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginLeft': '10px'})
        ], style={'display': 'flex', 'marginBottom': '10px'}),

        html.Div([
            html.Div([
                html.H4('GroupA_Detector2', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
                daq.BooleanSwitch(
                    id='groupA_dect2_switch',
                    on=False,
                    style={'transform': 'scale(1.1)'}
                )
            ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginRight': '10px'}),

            html.Div([
                html.H4('GroupB_Detector2', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
                daq.BooleanSwitch(
                    id='groupB_dect2_switch',
                    on=False,
                    style={'transform': 'scale(1.1)'}
                )
            ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginLeft': '10px'})
        ], style={'display': 'flex', 'marginBottom': '10px'}),

        html.Div([
            html.Div([
                html.H4('GroupA_Detector3', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
                daq.BooleanSwitch(
                    id='groupA_dect3_switch',
                    on=False,
                    style={'transform': 'scale(1.1)'}
                )
            ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginRight': '10px'}),

            html.Div([
                html.H4('GroupB_Detector3', style={'fontSize': '20px', 'color': '#003f5c', "font-weight": "100"}),
                daq.BooleanSwitch(
                    id='groupB_dect3_switch',
                    on=False,
                    style={'transform': 'scale(1.1)'}
                )
            ], style={'display': 'flex', 'alignItems': 'center', 'flex': '1', 'marginLeft': '10px'})
        ], style={'display': 'flex', 'marginBottom': '10px'})
    ]),
    dbc.Button('Check Raw Data Quality', id='check-data-quality-btn', color='primary', style={
        'padding': '15px', 'width': '100%', 'margin': '15px 0', 'fontSize': '22px'}),
    html.Div(id='check-raw-data-quality-desc', children='Check data quality', style={
        'fontFamily': 'Courier New',
        'fontSize': '20px',
        'textAlign': 'center',
        'marginTop': '15px',
        'color': '#2B2D42'
    })
]),
]),
]),
]),
                # Data Clean Tab
                dcc.Tab(label='Data Clean', children=[
                    html.Div([
                        html.H3('Data Cleaning', style={
                            'background': '#ed6a28',
                            'padding': '12px',
                            'textAlign': 'center',
                            'color': '#ECF0F1',
                            'fontWeight': 'bold',
                            'fontSize': '38px',
                            'borderRadius': '8px',
                        }),
                        html.Br(),
                        html.Div(
                                            style={
                                                'background': '#ffffff',
                                                'padding': '20px',
                                                'boxShadow': '0 4px 10px rgba(0, 0, 0, 0.5)',
                                                'marginBottom': '20px'
                                            },
                                            children=[
                                                html.Br(),
                                                # Subtract Dark Section
                html.Br(),
                html.Div('Subtract Dark', style={
                    'background': '#003f5c',
                    'padding': '10px',
                    'textAlign': 'center',
                    'color': 'white',
                    'fontWeight': 'bold',
                    'fontSize': '30px'
                }),
                html.Br(),
                html.Div(
                    children=[
                        dcc.Checklist(
                            id='preprocessing-options-subtract-dark',
                            options=[{'label': 'Subtract Dark', 'value': 'subtract-dark'}],
                            style={'textAlign': 'Center', 'fontSize': '28px', 'marginBottom': '10px', 'color': '#003f5c'},
                            inputStyle={'transform': 'scale(1.5)', 'marginRight': '10px'}
                        ),
                        html.Div('Subtract Noise', style={'fontFamily': 'Courier New', 'fontSize': '16px', 'textAlign': 'center', 'color': '#2B2D42'}),
                        html.Br()
                    ]
                ),
                html.Br(),
                        html.Br(),
                                                html.Div('Median Filtering', style={
                                                    'background': '#003f5c',
                                                    'padding': '10px',
                                                    'textAlign': 'center',
                                                    'color': 'white',
                                                    'fontWeight': 'bold',
                                                    'fontSize': '30px'
                                                }),
                                                html.Br(),
                                                html.Div(
                                                    children=[
                                                        dcc.Checklist(
                                                            id='preprocessing-options-median',
                                                            options=[
                                                                {'label': 'Median Filtering', 'value': 'median'},
                                                            ],
                                                            className='alignment-settings-section',
                                                            style={
                                                                'textAlign': 'Center',
                                                                'fontSize': '28px',
                                                                'marginBottom': '10px',
                                                                'color': '#003f5c',
                                                                "font-weight": "100",
                                                                'alignItems': 'center'
                                                            },
                                                            inputStyle={
                                                                'transform': 'scale(1.5)',  # Adjust this value to make the checkbox larger or smaller
                                                                'marginRight': '10px',  # Optional: add space between the checkbox and the label
                                                            }
                                                        ),
                                                        html.Br(),
                                                        html.Div(
                                                            style={'display': 'flex', 'alignItems': 'center', 'justifyContent': 'space-between'},
                                                            children=[
                                                                html.Div(
                                                                    className='app-controls-name',
                                                                    children='Filter Size',
                                                                    style={
                                                                        'textAlign': 'left',
                                                                        'fontSize': '26px',
                                                                        'marginBottom': '10px',
                                                                        'color': '#003f5c',
                                                                        "font-weight": "100",
                                                                        'marginLeft': '30px'
                                                                    }
                                                                ),
                                                                dcc.Input(
                                                    id='median-filter-size-input',
                                                    type='number',
                                                    min=1,
                                                    max=100,
                                                    step=1,
                                                    value=1,
                                                    style={
                                                        'width': '30%',
                                                        'padding': '5px',
                                                        'height': '24px',  # Increased height to fit the larger font
                                                        'textAlign': 'center',
                                                        'borderColor': '#003f5c',
                                                        'marginRight': '30px',
                                                        'fontSize': '24px'  # Increased font size for the number
                                                        }
                                                        )]
                                                        ),
                                                        html.Div(
                                                            children='Set Filter Size for Median Filtering',
                                                            style={
                                                                'fontFamily': 'Courier New',
                                                                'fontSize': '16px',
                                                                'textAlign': 'center',
                                                                'marginTop': '10px',
                                                                'color': '#2B2D42'
                                                            }
                                                        ),
                                                        html.Br(),
                    ]),
                ]),
                ]),
                ]),
                # Data Analysis Tab
                dcc.Tab(label='Data Analysis', children=[
                    html.Div([
                        html.H3('Data Analysis', style={
                            'background': '#ed6a28',
                            'padding': '12px',
                            'textAlign': 'center',
                            'color': '#ECF0F1',
                            'fontWeight': 'bold',
                            'fontSize': '38px',
                            'borderRadius': '8px',
                        }),
                        html.P("This tab will contain data analysis functionalities."),
                    ]),
                ]),
                dcc.Tab(label='Concentrations', children=[
        html.Div([
            html.H3('Concentrations', style={
                'background': '#ed6a28',
                'padding': '12px',
                'textAlign': 'center',
                'color': '#ECF0F1',
                'fontWeight': 'bold',
                'fontSize': '38px',
                'borderRadius': '8px',
            }),
            # Add the "Calculate Concentrations" button here
            html.Button('Calculate Concentrations', id='calculate-concentrations-btn', n_clicks=0, style={
                                'width': '95%', 'height': '70px', 'lineHeight': '70px',
                                'borderWidth': '1px', 'borderStyle': 'dashed', 'borderRadius': '5px',
                                'textAlign': 'center', 'margin': '15px', 'fontSize': '24px'}),
            html.Br(),
            html.Button('Download Concentrations Excel File', id='download_concentrations_excel_btn',  n_clicks=0, style={
                                'width': '95%', 'height': '70px', 'lineHeight': '70px',
                                'borderWidth': '1px', 'borderStyle': 'dashed', 'borderRadius': '5px',
                                'textAlign': 'center', 'margin': '15px', 'fontSize': '24px'}),
        html.Div(id='tabs-container'),
        ]),
    ]),
]),
        ], style={'width': '25%', 'padding': '15px', 'boxSizing': 'border-box'}),
        # Right side (3/4th width for Plot Section)
        html.Div([
            html.H3('Plot Section', style={
                'background': '#ed6a28',
                'padding': '12px',
                'textAlign': 'center',
                'color': '#ECF0F1',
                'fontWeight': 'bold',
                'fontSize': '38px',
                'borderRadius': '8px',
            }),
            dcc.Tabs(id='tabs', children=[
                dcc.Tab(label='Intensity vs Time', children=[
                    html.Div(id='intensity-time-plot'),
                ]),
            ]),
        ], style={'width': '75%', 'padding': '10px', 'boxSizing': 'border-box', 'borderLeft': '2px solid #3498DB'})
    ], style={'display': 'flex', 'height': '100vh'}),
])


# Define the output directory
output_dir = 'src/output_files/'

# Create the directory if it doesn't exist
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

def read_group_data(filepath):
    try:
        # Use pandas to read the Excel file
        data = pd.read_excel(filepath, engine="openpyxl")
        data_df = pd.DataFrame(data)
        # Extract the data starting from row 9
        data = data_df.iloc[9:, :]
        data.columns = ['Time', 'System Time (s)', 'Sample Time (s)',
                    'LED_A_784_DET1', 'LED_A_784_DET2', 'LED_A_784_DET3', 
                    'LED_A_800_DET1', 'LED_A_800_DET2', 'LED_A_800_DET3', 
                    'LED_A_818_DET1', 'LED_A_818_DET2', 'LED_A_818_DET3', 
                    'LED_A_835_DET1', 'LED_A_835_DET2', 'LED_A_835_DET3',
                    'LED_A_851_DET1', 'LED_A_851_DET2', 'LED_A_851_DET3', 
                    'LED_A_881_DET1', 'LED_A_881_DET2', 'LED_A_881_DET3', 
                    'LED_A_DARK_DET1', 'LED_A_DARK_DET2', 'LED_A_DARK_DET3',
                    'LED_B_784_DET1', 'LED_B_784_DET2', 'LED_B_784_DET3', 
                    'LED_B_800_DET1', 'LED_B_800_DET2', 'LED_B_800_DET3', 
                    'LED_B_818_DET1', 'LED_B_818_DET2', 'LED_B_818_DET3', 
                    'LED_B_835_DET1', 'LED_B_835_DET2', 'LED_B_835_DET3',
                    'LED_B_851_DET1', 'LED_B_851_DET2', 'LED_B_851_DET3', 
                    'LED_B_881_DET1', 'LED_B_881_DET2', 'LED_B_881_DET3', 
                    'LED_B_DARK_DET1', 'LED_B_DARK_DET2', 'LED_B_DARK_DET3']
        Time = data['Time']
        
        # Group mappings
        GroupA_Detector1 = ['LED_A_784_DET1', 'LED_A_800_DET1', 'LED_A_818_DET1', 'LED_A_835_DET1', 'LED_A_851_DET1', 'LED_A_881_DET1', 'LED_A_DARK_DET1']
        GroupA_Detector2 = ['LED_A_784_DET2', 'LED_A_800_DET2', 'LED_A_818_DET2', 'LED_A_835_DET2', 'LED_A_851_DET2', 'LED_A_881_DET2', 'LED_A_DARK_DET2']
        GroupA_Detector3 = ['LED_A_784_DET3', 'LED_A_800_DET3', 'LED_A_818_DET3', 'LED_A_835_DET3', 'LED_A_851_DET3', 'LED_A_881_DET3', 'LED_A_DARK_DET3']
    
        GroupB_Detector1 = ['LED_B_784_DET1', 'LED_B_800_DET1', 'LED_B_818_DET1', 'LED_B_835_DET1', 'LED_B_851_DET1', 'LED_B_881_DET1', 'LED_B_DARK_DET1']
        GroupB_Detector2 = ['LED_B_784_DET2', 'LED_B_800_DET2', 'LED_B_818_DET2', 'LED_B_835_DET2', 'LED_B_851_DET2', 'LED_B_881_DET2', 'LED_B_DARK_DET2']
        GroupB_Detector3 = ['LED_B_784_DET3', 'LED_B_800_DET3', 'LED_B_818_DET3', 'LED_B_835_DET3', 'LED_B_851_DET3', 'LED_B_881_DET3', 'LED_B_DARK_DET3']
    
        return data, Time, GroupA_Detector1, GroupA_Detector2, GroupA_Detector3, GroupB_Detector1, GroupB_Detector2, GroupB_Detector3
    except Exception as e:
        print(f"Error reading the file: {e}")
        return None, None, None, None, None, None, None, None

def resample_data(df, option):
    if option == 'average':
        resampled_df = df.resample('1S', on='Time').mean().reset_index()
    elif option == 'accumulation':
        resampled_df = df.resample('1S', on='Time').sum().reset_index()

    # Convert Time from milliseconds to seconds (resample already adjusts)
    resampled_df['Time'] = (resampled_df['Time'] - resampled_df['Time'].min()).dt.total_seconds()
    print('resampled_df[Time]', resampled_df['Time'])
    return resampled_df



@callback(
    [Output('store-file-path', 'data'),
     Output('file-names', 'children')],
    [Input('upload-data', 'filename'),
     Input('upload-data', 'contents'),
     Input('resample-option', 'value')],
    prevent_initial_call=True
)
def upload_and_resample_data(filename, contents, selected_option):
    if filename and contents:
        # Decode the uploaded file's contents
        content_type, content_string = contents.split(',')
        decoded = base64.b64decode(content_string)

        # Prepare the file path and save the file
        directory = "src/uploads"  # Using relative path
        files = glob.glob(os.path.join(directory, "*"))
        for f in files:
            os.remove(f)  # Clear old files before saving the new one

        filepath = os.path.join(directory, filename)

        with open(filepath, 'wb') as f:
            f.write(decoded)
        
        # Read the file (assuming the function `read_group_data` is defined elsewhere)
        voltages, Time, GroupA_Detector1, GroupA_Detector2, GroupA_Detector3, GroupB_Detector1, GroupB_Detector2, GroupB_Detector3 = read_group_data(filepath)
        
        # Convert the voltages dataframe to a DataFrame for resampling
        voltages_df = pd.DataFrame(voltages)
        
        # Ensure 'Time' column exists and is in Time format
        voltages_df['Time'] = pd.to_datetime(voltages_df['Time'])

        # Apply the resampling based on the selected option
        if selected_option:
            resampled_df = resample_data(voltages_df, selected_option)
        else:
            resampled_df = voltages_df  # No resampling if no option is selected

        # Create a new directory for resampled data if it doesn't exist
        resampled_directory = 'src/resampled_data'
        os.makedirs(resampled_directory, exist_ok=True)

        # Remove old files in the resampled folder
        files = glob.glob(os.path.join(resampled_directory, "*"))
        for f in files:
            os.remove(f)

        # Prepare the file name and path for the resampled data (as Excel)
        resampled_filename = filename.replace('.csv', '_resampled_data.xlsx')
        resampled_filepath = os.path.join(resampled_directory, resampled_filename)

        # Save the resampled data to Excel
        resampled_df.to_excel(resampled_filepath, index=False)

        # Return the path of the resampled file and update the UI with the file name
        return resampled_filepath, f'File Uploaded and Saved as Excel: {resampled_filename}'

    return None, 'No file uploaded yet'



@callback(
    Output('download-file-snirf', 'data'),  # Make sure 'data' is the output, not 'contents'
    Input('btn_rawdata_snirf', 'n_clicks'),
    Input('upload-data', 'filename'),
    prevent_initial_call=True
)
def rawdata_snirf_download(n_clicks, filename):
    if n_clicks and filename:
        print('Button clicked to download')
        
        # Generate SNIRF file
        directory = "src/uploads"  # Path for file
        filepath = os.path.join(directory, filename)
        print('filepath', filepath)

        # Assuming snirf_file is a function that returns the file in the required format
        snirf_file, snirf_file_name = create_snirf(filepath)  # Replace with actual file creation logic
        return dcc.send_file(snirf_file)
    return None

# Function to create the intensity figure
def create_intensity_figure(voltages, spectra_list, title, time_unit='ms'):
    data = []
    y_max = 0  # Variable to track the maximum y value across all spectra
    for spectrum in spectra_list:
        if spectrum in voltages.columns:
            y = voltages[spectrum].values
            y_max = max(y_max, max(y, default=0))  # Update y_max if current y has a higher value
            x = np.arange(len(y))
            trace = go.Scatter(x=x, y=y, mode='lines', name=spectrum)
            data.append(trace)
    
    layout = go.Layout(
        title={'text': title, 'font_size': 24},  # Title font size
        xaxis={'title': f'Time ({time_unit})', 'range': [0, len(x)], 'title_font_size': 20, 'tickfont_size': 20},  # Axis labels and ticks font size
        yaxis={'title': 'Voltage (V)', 'range': [0, y_max], 'title_font_size': 20, 'tickfont_size': 20},  # Axis labels and ticks font size
        legend={'font_size': 16},  # Legend font size
        height=1300
    )
    
    return go.Figure(data=data, layout=layout)



def concentrations(filename):
    # Define the paths to the resampled and uploads directories
    resampled_directory = 'src/resampled_data'
    uploads_directory = 'src/uploads'
    
    # Construct the file paths
    resampled_filepath = os.path.join(resampled_directory, filename)
    uploaded_filepath = os.path.join(uploads_directory, filename)
    
    # Check if the file exists in the resampled data directory
    if os.path.exists(resampled_filepath):
        print(f"Opening resampled file: {filename}")
        data = pd.read_excel(resampled_filepath)
        
    # If file does not exist in resampled, check the uploads directory
    elif os.path.exists(uploaded_filepath):
        print(f"Opening uploaded raw data file: {filename}")
        data = pd.read_csv(uploaded_filepath)  # Assuming the raw data is in CSV format
        
    else:
        print(f"File {filename} not found in either directory.")
        return None  # If the file is not found, return None

    # Initialize PhantomDataProcessor instance
    processor = PhantomDataProcessor()

    # Call formula to extract necessary data
    group_a_1, group_a_2, group_a_3, group_b_1, group_b_2, group_b_3, wavelengths, dpf, wavelength_dependency, ext_coeffs_inv = processor.formula(data)

    # Call UCLN to get the wavelength-dependent attenuation values and concentrations
    conc_a_1_df, conc_a_2_df, conc_a_3_df, conc_b_1_df, conc_b_2_df, conc_b_3_df, ext_coeffs_inv, atten_a_1, atten_a_2, atten_a_3, atten_b_1, atten_b_2, atten_b_3, ext_coeffs_inv = processor.UCLN(data)

    df_sto2_A, df_sto2_B =  SRS(atten_a_1, atten_a_2, atten_a_3, atten_b_1, atten_b_2, atten_b_3, ext_coeffs_inv)
    wavelength = [784, 800, 818, 835, 851, 881]
    LED_A_det_seps = [3, 5]  # for LED_A_DET1 and LED_A_DET3
    LED_B_det_seps = [5, 3]  # for LED_B_DET1 and LED_B_DET3
    print('data head', data.head(4))
    print('data columns', data.columns)
    sto2_dual_slope = dual_slope_wavelength(data, wavelength, LED_A_det_seps, LED_B_det_seps, dsf=8)
    
    df_sto2_A = pd.DataFrame(df_sto2_A, columns=["Sto2_A"]) 
    df_sto2_B = pd.DataFrame(df_sto2_B, columns=["Sto2_B"]) 
    df_sto2_dual_slope = pd.DataFrame(sto2_dual_slope, columns=["dual_slope_sto2"]) 
    
    # Create a new directory for concentrations data if it doesn't exist
    concentrations_directory = 'src/concentration_data'
    os.makedirs(concentrations_directory, exist_ok=True)

    # Remove old files in the concentrations folder
    files = glob.glob(os.path.join(concentrations_directory, "*"))
    for f in files:
        os.remove(f)

    # Create Excel file with concentrations data
    wb = openpyxl.Workbook()
    default_sheet =  wb.active
    wb.remove(default_sheet)


    # Add the individual DataFrames as sheets
    wb.create_sheet(title="Concentration LED A-DET 1")
    for r in dataframe_to_rows(conc_a_1_df, index=False, header=True):
        wb["Concentration LED A-DET 1"].append(r)

    wb.create_sheet(title="Concentration LED A-DET 2")
    for r in dataframe_to_rows(conc_a_2_df, index=False, header=True):
        wb["Concentration LED A-DET 2"].append(r)

    wb.create_sheet(title="Concentration LED A-DET 3")
    for r in dataframe_to_rows(conc_a_3_df, index=False, header=True):
        wb["Concentration LED A-DET 3"].append(r)

    wb.create_sheet(title="Concentration LED B-DET 1")
    for r in dataframe_to_rows(conc_b_1_df, index=False, header=True):
        wb["Concentration LED B-DET 1"].append(r)

    wb.create_sheet(title="Concentration LED B-DET 2")
    for r in dataframe_to_rows(conc_b_2_df, index=False, header=True):
        wb["Concentration LED B-DET 2"].append(r)

    wb.create_sheet(title="Concentration LED B-DET 3")
    for r in dataframe_to_rows(conc_b_3_df, index=False, header=True):
        wb["Concentration LED B-DET 3"].append(r)

    wb.create_sheet(title="Tissue oxygen saturation(StO2) LED A")
    # Append the rows of df_sto2_A to the newly created sheet
    for r in dataframe_to_rows(df_sto2_A, index=False, header=True):
        wb["Tissue oxygen saturation(StO2) LED A"].append(r)

    # Create a new sheet with the title "df_sto2_B"
    wb.create_sheet(title="Tissue oxygen saturation(StO2) LED B")
    # Append the rows of df_sto2_B to the newly created sheet
    for r in dataframe_to_rows(df_sto2_B, index=False, header=True):
        wb["Tissue oxygen saturation(StO2) LED B"].append(r)

    # Create a new sheet with the title "df_sto2_B"
    wb.create_sheet(title="Tissue oxygen saturation(StO2) Dual Slope")
    # Append the rows of dual_slope_sto2 to the newly created sheet
    for r in dataframe_to_rows(df_sto2_dual_slope, index=False, header=True):
        wb["Tissue oxygen saturation(StO2) Dual Slope"].append(r)

    # Save the workbook
    wb.save("src/concentration_data/concentrations.xlsx")

    # Return necessary results
    return {
        'Concentration LED A-DET 1': conc_a_1_df,
        'Concentration LED A-DET 2': conc_a_2_df,
        'Concentration LED A-DET 3': conc_a_3_df,
        'Concentration LED B-DET 1': conc_b_1_df,
        'Concentration LED B-DET 2': conc_b_2_df,
        'Concentration LED B-DET 3': conc_b_3_df,
        'Tissue oxygen saturation(StO2) LED A': df_sto2_A,
        'Tissue oxygen saturation(StO2) LED B': df_sto2_B,
        'Tissue oxygen saturation(StO2) Dual Slope': df_sto2_dual_slope
    }

def load_and_plot_concentrations():
    # Path to the concentrations data folder
    concentrations_directory = 'src/concentration_data'
    concentrations_file = os.path.join(concentrations_directory, 'concentrations.xlsx')
    
    # Check if the file exists
    if not os.path.exists(concentrations_file):
        print(f"File not found: {concentrations_file}")
        return []

    # Load the Excel file with pandas
    excel_data = pd.ExcelFile(concentrations_file)
    
    tabs = []
    
    # Loop through each sheet in the Excel file
    for sheet_name in excel_data.sheet_names:
        df = pd.read_excel(excel_data, sheet_name=sheet_name)
        
        # Create the figure for the plot
        fig = px.line(df, title=sheet_name)
        
        # Customizing axis labels based on the sheet name
        if "Concentration" in sheet_name:
            fig.update_layout(
                xaxis_title="Time",  # X axis is "Time" for concentration plots
                yaxis_title="ΔC (mM)"  # Y axis is "ΔC (mM)" for concentration plots
            )
        elif "Tissue oxygen saturation" in sheet_name:
            fig.update_layout(
                xaxis_title="Time",  # X axis is "Time" for StO2 plots
                yaxis_title="Saturation StO2 (%)"  # Y axis is "Saturation StO2 (%)" for StO2 plots
            )

        # Create a tab for each sheet with a plot
        tabs.append(
            dcc.Tab(
                label=sheet_name,  # Tab label is the sheet name
                value=sheet_name,  # Tab value is the sheet name
                children=[
                    html.Div(
                        children=[
                            html.H4(f"{sheet_name}"),
                            dcc.Graph(figure=fig, style={
                                'height': '70vh',  # 50% of the viewport height
                                'width': '100%',
                            }),  # Plot the data here
                        ]
                    )
                ]
            )
        )
    
    return tabs

def data_quality_check(voltages):
    # Define the signal-dark detector mapping (you can adjust the keys and values as necessary)
    signal_dark_dictionary = {
        'LED_A_DET1': 'LED_A_DARK_DET1',
        'LED_A_DET2': 'LED_A_DARK_DET2',
        'LED_A_DET3': 'LED_A_DARK_DET3',
        'LED_B_DET1': 'LED_B_DARK_DET1',
        'LED_B_DET2': 'LED_B_DARK_DET2',
        'LED_B_DET3': 'LED_B_DARK_DET3',
    }
    
    # Function to calculate SNR
    def calculate_snr(signal_data, dark_data):
        signal = np.mean(signal_data)
        dark_mean = np.mean(dark_data)
        noise = np.std(dark_data)
        snr = (signal - dark_mean) / noise
        return snr

    # SNR Histogram: Visualizing the distribution of the Signal-to-Noise Ratios
    snr_histogram = go.Figure()
    snr_values = []
    
    for signal_col, dark_col in signal_dark_dictionary.items():
        snr_value = calculate_snr(voltages[signal_col], voltages[dark_col])
        snr_values.append(snr_value)

    snr_histogram.add_trace(go.Histogram(
        x=snr_values,  # SNR values
        nbinsx=20, 
        marker_color='skyblue', 
        name="SNR Histogram"
    ))
    snr_histogram.update_layout(
        title="Histogram of SNR",
        xaxis_title="SNR",
        yaxis_title="Frequency",
        template="plotly_dark"
    )
    
    # SNR Bar Chart: Visualizing the average SNR for each group
    avg_snr = [np.mean(voltages[key].values) if isinstance(voltages[key], pd.DataFrame) else np.mean(voltages[key]) for key in voltages.keys()]
    snr_bar_chart = go.Figure()
    snr_bar_chart.add_trace(go.Bar(
        x=list(voltages.keys()),  # Use detector names as x-axis labels
        y=avg_snr, 
        marker_color='royalblue',
        name="Average SNR"
    ))
    snr_bar_chart.update_layout(
        title="Average SNR for Each Group",
        xaxis_title="Group",
        yaxis_title="SNR",
        template="plotly_dark"
    )

    # Calculate the NEP for each dark detector
    nep_values = [np.std(voltages.get(detector, pd.DataFrame()).values) for detector in [
            'LED_A_DARK_DET1', 'LED_A_DARK_DET2', 'LED_A_DARK_DET3',
            'LED_B_DARK_DET1', 'LED_B_DARK_DET2', 'LED_B_DARK_DET3'
        ]]
    
    nep_group_names = [
        'GroupA_Detector1', 'GroupA_Detector2', 'GroupA_Detector3',
        'GroupB_Detector1', 'GroupB_Detector2', 'GroupB_Detector3'
    ]
    nep_fig = go.Figure()
    nep_fig.add_trace(go.Bar(
        x=nep_group_names,
        y=nep_values,
        text=nep_values,  # Display NEP values above each bar
        textposition='outside',
        marker=dict(color='royalblue')
    ))
    nep_fig.update_layout(
        title="NEP Values for Each Group",
        xaxis_title="Group",
        yaxis_title="NEP Value",
        template="plotly_dark",
        showlegend=False
    )

    return snr_histogram, snr_bar_chart, nep_fig

@callback(
    Output('tabs', 'children'),
    Input('view-graph-btn', 'n_clicks'),
    Input('check-data-quality-btn', 'n_clicks'),
    Input('calculate-concentrations-btn', 'n_clicks'), 
    Input('download_concentrations_excel_btn', 'n_clicks'),
    State('groupA_dect1_spectras', 'on'),
    State('groupA_dect2_spectras', 'on'),
    State('groupB_dect1_spectras', 'on'),
    State('groupB_dect2_spectras', 'on'),
    State('select_all_switch', 'on'),
    State('intensities-options-dropdown', 'value'),
    State('store-file-path', 'data'),
    prevent_initial_call=True
)
def update_tabs_and_plots(view_intensity_n_clicks, check_data_quality_n_clicks, calculate_concentrations_n_clicks,download_concentrations_n_clicks, groupA1_on, groupA2_on, groupB1_on, groupB2_on, select_all_on, selected_spectra, filepath):
    ctx = dash.callback_context
    if not ctx.triggered:
        return []

    button_id = ctx.triggered[0]['prop_id'].split('.')[0]
    print(f"Triggered by {button_id}")
    
    # Handling for "View Graph" button (existing logic)
    if button_id == 'view-graph-btn' and filepath:
        voltages, Time, GroupA_Detector1, GroupA_Detector2, GroupA_Detector3, GroupB_Detector1, GroupB_Detector2, GroupB_Detector3 = read_group_data(filepath)
        
        tabs = []
        if select_all_on:
            groupA1_on = groupA2_on = groupA3_on = True
            groupB1_on = groupB2_on = groupB3_on = True

        if groupA1_on:
            tabs.append(dcc.Tab(label='GroupA_Detector1', value='GroupA_Detector1', children=[
                dcc.Graph(figure=create_intensity_figure(voltages, GroupA_Detector1, 'GroupA_Detector1'))
            ]))
        if groupA2_on:
            tabs.append(dcc.Tab(label='GroupA_Detector2', value='GroupA_Detector2', children=[
                dcc.Graph(figure=create_intensity_figure(voltages, GroupA_Detector2, 'GroupA_Detector2'))
            ]))
        if groupA3_on:
            tabs.append(dcc.Tab(label='GroupA_Detector3', value='GroupA_Detector3', children=[
                dcc.Graph(figure=create_intensity_figure(voltages, GroupA_Detector3, 'GroupA_Detector3'))
            ]))
        if groupB1_on:
            tabs.append(dcc.Tab(label='GroupB_Detector1', value='GroupB_Detector1', children=[
                dcc.Graph(figure=create_intensity_figure(voltages, GroupB_Detector1, 'GroupB_Detector1'))
            ]))
        if groupB2_on:
            tabs.append(dcc.Tab(label='GroupB_Detector2', value='GroupB_Detector2', children=[
                dcc.Graph(figure=create_intensity_figure(voltages, GroupB_Detector2, 'GroupB_Detector2'))
            ]))
        if groupB3_on:
            tabs.append(dcc.Tab(label='GroupB_Detector3', value='GroupB_Detector3', children=[
                dcc.Graph(figure=create_intensity_figure(voltages, GroupB_Detector3, 'GroupB_Detector3'))
            ]))

        if selected_spectra:
            tabs.append(dcc.Tab(label='Selected Intensities', value='Selected_Intensities', children=[
                dcc.Graph(figure=create_intensity_figure(voltages, selected_spectra, 'Selected Intensities'))
            ]))

        return tabs

    elif check_data_quality_n_clicks and filepath:
        voltages, Time, GroupA_Detector1, GroupA_Detector2, GroupA_Detector3, GroupB_Detector1, GroupB_Detector2, GroupB_Detector3 = read_group_data(filepath)

        # Call the data_quality_check function to get the figures
        snr_histogram, snr_bar_chart, nep_fig = data_quality_check(voltages)

        # Create the tab layout for data quality check
        return [
            dcc.Tab(
                label='Check Raw Data Quality',
                value='check_raw_data_quality',
                children=[
                    html.Div(
                        children=[
                            html.Div(
                                dcc.Graph(figure=snr_histogram),
                                style={'width': '100%', 'display': 'inline-block', 'margin-bottom': '20px'}
                            ),
                            html.Div(
                                dcc.Graph(figure=snr_bar_chart),
                                style={'width': '100%', 'display': 'inline-block'}
                            ),
                            html.Div(
                                dcc.Graph(figure=nep_fig),
                                style={'width': '100%', 'display': 'inline-block'}
                            )
                        ]
                    )
                ]
            )
        ]


    # Handling for "Calculate Concentrations" button
    elif calculate_concentrations_n_clicks and filepath:
        print("Triggered by Calculate Concentrations button")

        # Assuming the concentrations function is implemented
        concentrations(filepath)
        
        # Return tabs related to concentrations
        tabs = load_and_plot_concentrations()
        
        return tabs

    # Handling for "Download Concentrations Excel File" button
    elif button_id == 'download_concentrations_excel_btn':
        csv_dir = '/home/darshana/Desktop/FSM_v2-dashboard/src/concentration_data'

        # List all .xlsx files in the directory
        files = [file for file in os.listdir(csv_dir) if file.endswith('.xlsx')]

        if files:
            file_path = os.path.join(csv_dir, files[0])

            if os.path.exists(file_path):
                with open(file_path, 'rb') as f:
                    file_data = f.read()

                return dcc.send_bytes(file_data, filename=files[0])

    return []  # Default return if no button is triggered


if __name__ == '__main__':
    app.run_server(debug=True) 
